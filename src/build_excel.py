"""
build_excel.py
--------------
Convert the JSON output of `extract_cdx.py` into a formatted Excel spreadsheet.

The script dynamically detects all annotation columns present in the data
(MW, CLogP, SeqCount_*, etc.) and includes them as additional columns after
the mandatory `Compound_ID` and `SMILES` columns.

Usage
-----
    python src/build_excel.py output/structures.json -o output/structures.xlsx
"""

import json
import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from extract_cdx import _ID_ANNOTATION_KEYS


# ── Styling constants ─────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

ALT_FILL   = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Column widths (characters)
COL_WIDTHS = {
    'Compound_ID': 45,
    'SMILES':      80,
    'Stereo_Note': 18,
    'MW':          10,
    'CLogP':       10,
    'ALogP':       10,
    'Source_File': 22,
}
SEQCOUNT_WIDTH = 18
DEFAULT_WIDTH  = 15


# ── Core logic ────────────────────────────────────────────────────────────────

def load_results(json_path: str) -> list[dict]:
    with open(json_path, encoding='utf-8') as f:
        return json.load(f)['results']


def _normalise_annotation_keys(annotations: dict) -> dict:
    """
    Normalise annotation keys so that different naming conventions used by
    different ChemDraw CDX export versions are unified under one name.

    Specifically, ``SeqCount_*`` keys (produced by some DEL export templates)
    are renamed to ``SequenceCount_*`` to match the canonical column name.
    This prevents the same data from appearing in two separate columns when
    rows from different CDX layout variants are combined in one spreadsheet.

    Examples
    --------
    ``SeqCount_S1_1h``  → ``SequenceCount_S1_1h``
    ``SeqCount_S01``    → ``SequenceCount_S01``
    """
    normalised: dict = {}
    for k, v in annotations.items():
        if k.startswith('SeqCount_'):
            new_key = 'SequenceCount_' + k[len('SeqCount_'):]
        else:
            new_key = k
        # If the canonical key already exists (shouldn't happen, but be safe),
        # prefer the non-empty value.
        if new_key in normalised:
            normalised[new_key] = normalised[new_key] or v
        else:
            normalised[new_key] = v
    return normalised


def build_dataframe(results: list[dict]) -> pd.DataFrame:
    """
    Flatten the nested JSON result list into a pandas DataFrame.

    Mandatory columns: Compound_ID, SMILES
    Optional columns:  MW, CLogP, all annotation keys, Source_File

    Annotation keys are normalised before building the DataFrame so that
    ``SeqCount_*`` and ``SequenceCount_*`` variants are merged into a single
    ``SequenceCount_*`` column.
    """
    # Normalise annotation keys in-place for all records
    for r in results:
        if 'annotations' in r:
            r['annotations'] = _normalise_annotation_keys(r['annotations'])

    # Discover all annotation keys across all records, preserving the
    # first-seen insertion order — this matches the top-to-bottom order
    # of annotations in the source CDX objects, which becomes left-to-right
    # in the Excel output.
    all_ann_keys: list[str] = []
    seen: set[str] = set()

    for r in results:
        for k in r.get('annotations', {}):
            if k not in seen:
                seen.add(k)
                all_ann_keys.append(k)

    # Preserve CDX insertion order for all annotation columns.
    # No alphabetical sort — the order reflects the original CDX annotation
    # sequence, which is meaningful (e.g. MW, CLogP, SeqCounts in assay order).
    rows = []
    for r in results:
        ann = r.get('annotations', {})

        # compound_id from extract_cdx is authoritative (handles all CDX layouts).
        # Annotation-key fallback is a safety net for hand-edited JSON.
        compound_id = r.get('compound_id')
        if not compound_id:
            for kw in _ID_ANNOTATION_KEYS:
                compound_id = ann.get(kw)
                if compound_id:
                    break
        if not compound_id:
            compound_id = f"Unknown_{r['index']}"
        if compound_id in ('N', ''):
            compound_id = f"Unknown_{r['index']}"

        row: dict = {
            'Compound_ID': compound_id,
            'SMILES':      r.get('smiles', ''),
            'Stereo_Note': r.get('stereo_note', ''),
        }
        for k in all_ann_keys:
            row[k] = ann.get(k, '')
        row['Source_File'] = r.get('file', '')

        rows.append(row)

    return pd.DataFrame(rows)


def apply_formatting(ws, df: pd.DataFrame):
    """Apply header styling, alternating row colours, borders, and column widths."""
    # Header row
    for cell in ws[1]:
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for cell in row:
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=False)

    # Column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_name in COL_WIDTHS:
            ws.column_dimensions[letter].width = COL_WIDTHS[col_name]
        elif 'SeqCount' in col_name or 'SequenceCount' in col_name:
            ws.column_dimensions[letter].width = SEQCOUNT_WIDTH
        else:
            ws.column_dimensions[letter].width = DEFAULT_WIDTH

    # Row heights
    ws.row_dimensions[1].height = 30
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 18

    # Freeze header and enable auto-filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def build_excel(json_path: str, output_path: str):
    """
    Main entry point: load JSON, build DataFrame, write formatted Excel.

    Parameters
    ----------
    json_path : str
        Path to the JSON file produced by extract_cdx.py.
    output_path : str
        Path to write the output .xlsx file.
    """
    results = load_results(json_path)
    df = build_dataframe(results)

    print(f"Rows:    {len(df)}")
    print(f"Columns: {list(df.columns)}")

    missing = df[df['SMILES'].isna() | (df['SMILES'] == '')]
    if len(missing):
        print(f"\n[WARN] {len(missing)} row(s) have no SMILES:")
        print(missing[['Compound_ID', 'Source_File']].to_string(index=False))

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Structures', index=False)
        apply_formatting(writer.sheets['Structures'], df)

    print(f"\nExcel saved → {output_path}")


# ── CLI entry point ───────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Build a formatted Excel file from extracted CDX structure data.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument('input_json',  help='Path to the JSON file from extract_cdx.py.')
    parser.add_argument(
        '--output', '-o',
        default='output/structures.xlsx',
        help='Path to the output Excel file (default: output/structures.xlsx).'
    )
    args = parser.parse_args()
    build_excel(args.input_json, args.output)
